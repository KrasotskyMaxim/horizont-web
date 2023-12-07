import aiohttp
from openpyxl import load_workbook

from rest_framework.exceptions import ValidationError

from api.constants import CARD_API_URL_TEMPLATE, Item


def get_articles_from_file(file):
    try:
        workbook = load_workbook(file, read_only=True)
        return [row[0] for row in workbook.active.iter_rows(max_col=1, values_only=True)]
    except Exception as e:
        raise ValidationError(f"Failed to read the file. {str(e)}")


async def get_article_data(article):
    async with aiohttp.ClientSession() as session:
        async with session.get(CARD_API_URL_TEMPLATE.format(article=article)) as response:
            data = await response.json()
            try:
                card_data = data["data"]["products"][0]
            except Exception as e:
                return Item(article=article)
            return Item(
                article=article,
                brand=card_data["brand"],
                title=card_data["name"],
            )
